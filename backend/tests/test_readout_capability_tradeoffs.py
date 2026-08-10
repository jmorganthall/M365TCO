"""Full reconciliation: a right-sizing move that sheds capability must SHOW the
trade-off everywhere the savings are shown — the /compute result, the HTML
readout (headline caveat + a Capability trade-offs section), and the xlsx — so
the saved-dollars figure is never presented as a free win.
"""

import io

from openpyxl import load_workbook


def _ems_drop_engagement(client, name):
    """Persona on Office 365 E3 + Enterprise Mobility + Security E3, targeted at
    just Office 365 E3 — the move drops EMS capability (MFA/CA, MDM)."""
    eid = client.post("/api/engagements", json={"customer_name": name}).json()["id"]
    p = client.post(f"/api/engagements/{eid}/personas",
                    json={"name": "Non-Store Users", "headcount": 700}).json()
    for sku in ("Office 365 E3", "Enterprise Mobility + Security E3"):
        client.post(f"/api/engagements/{eid}/current-licenses",
                    json={"sku_reference": sku, "quantity_assigned": 700,
                          "unit_price_paid_annual": 300, "persona_ids": [p["id"]]})
    client.post(f"/api/engagements/{eid}/scenarios",
                json={"persona_id": p["id"], "target_sku_reference": "Office 365 E3",
                      "target_unit_price_annual": 300, "in_scope": True})
    return eid, p["id"]


def test_compute_result_exposes_dropped_capability(client):
    eid, pid = _ems_drop_engagement(client, "Reconcile Compute Co")
    result = client.post(f"/api/engagements/{eid}/compute").json()
    dropped = result["dropped_capability"]
    entry = next(d for d in dropped if d["persona_id"] == pid)
    names = {o["name"] for o in entry["outcomes"]}
    assert "MFA & Conditional Access" in names
    assert "Mobile Device & App Management (MDM/MAM)" in names


def test_html_readout_shows_capability_tradeoffs_and_headline_caveat(client):
    eid, _ = _ems_drop_engagement(client, "Reconcile HTML Co")
    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "Capability trade-offs" in html
    assert "MFA &amp; Conditional Access" in html  # html-escaped
    assert "Mobile Device &amp; App Management (MDM/MAM)" in html
    # The headline caveat ties the number to the trade-off, so it isn't read alone.
    assert "capability trade-off" in html


def test_xlsx_has_a_capability_changes_sheet(client):
    eid, _ = _ems_drop_engagement(client, "Reconcile Xlsx Co")
    data = client.get(f"/api/engagements/{eid}/readout.xlsx").content
    wb = load_workbook(io.BytesIO(data))
    assert "Capability changes" in wb.sheetnames
    rows = list(wb["Capability changes"].iter_rows(values_only=True))
    dropped = {r[3] for r in rows[1:] if r[2] == "Dropped"}
    assert "MFA & Conditional Access" in dropped


def test_no_tradeoff_section_when_nothing_is_dropped(client):
    """A clean upgrade (Microsoft 365 E3 → Microsoft 365 E5) drops nothing, so the
    trade-off section and headline caveat are absent — no false alarm."""
    eid = client.post("/api/engagements", json={"customer_name": "Clean Upgrade Co"}).json()["id"]
    p = client.post(f"/api/engagements/{eid}/personas",
                    json={"name": "KW", "headcount": 100}).json()
    client.post(f"/api/engagements/{eid}/current-licenses",
                json={"sku_reference": "Microsoft 365 E3", "quantity_assigned": 100,
                      "unit_price_paid_annual": 300, "persona_ids": [p["id"]]})
    client.post(f"/api/engagements/{eid}/scenarios",
                json={"persona_id": p["id"], "target_sku_reference": "Microsoft 365 E5",
                      "target_unit_price_annual": 600, "in_scope": True})

    result = client.post(f"/api/engagements/{eid}/compute").json()
    assert result["dropped_capability"] == []
    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "Capability trade-offs" not in html
    assert "capability trade-off" not in html
