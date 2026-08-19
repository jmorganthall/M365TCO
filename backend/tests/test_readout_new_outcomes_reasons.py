"""An empty New-outcomes story must never be silent — or read as a failure.

"No new capabilities for this persona" has three very different causes — the
target maps to no capability at all, the current licensing was counted org-wide
because it carries no persona tag, or the target genuinely adds nothing. Omitting
the persona (the old behaviour) made all three look identical to a lost section,
and the first two also invented a full-capability "drop". Every surface that
shows the story must name the reason instead — and where the move is a
consolidation play, name what it DOES deliver (tools folded in, spend
right-sized, a smaller vendor/audit surface) rather than only what it does not.
"""

import io

from openpyxl import load_workbook


def _engagement(client, name, *, target, tag_licence=True):
    eid = client.post("/api/engagements", json={"customer_name": name}).json()["id"]
    kw = client.post(f"/api/engagements/{eid}/personas",
                     json={"name": "Knowledge Workers", "headcount": 1000}).json()
    client.post(f"/api/engagements/{eid}/current-licenses",
                json={"sku_reference": "Microsoft 365 E5", "quantity_assigned": 1000,
                      "unit_price_paid_annual": 690,
                      "persona_ids": [kw["id"]] if tag_licence else []})
    client.post(f"/api/engagements/{eid}/scenarios",
                json={"persona_id": kw["id"], "target_sku_reference": target,
                      "target_unit_price_annual": 690, "in_scope": True})
    return eid, kw["id"]


def _entry(result, pid):
    return next(n for n in result["new_outcomes"] if n["persona_id"] == pid)


def test_in_scope_persona_with_nothing_new_is_still_listed_with_a_reason(client):
    """Target = what they already hold, nothing consolidated, no spend change:
    say exactly that, without dressing it up."""
    eid, pid = _engagement(client, "Nothing New Co", target="Microsoft 365 E5")
    result = client.post(f"/api/engagements/{eid}/compute").json()
    entry = _entry(result, pid)
    assert entry["outcomes"] == []
    assert entry["empty_reason"] == "covered_today"
    assert "No new functional outcomes" in entry["empty_reason_text"]
    assert "no change in spend" in entry["empty_reason_text"]
    assert result["dropped_capability"] == []

    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "<h2>New outcomes</h2>" in html          # the section survives
    assert "Knowledge Workers" in html
    assert "No new functional outcomes" in html


def test_consolidation_play_is_told_as_consolidation_not_as_a_shortfall(client):
    """The common case: the target adds no NEW outcome but retires third-party
    tools and spend. The readout must lead with what that IS worth — tools folded
    in, a smaller vendor/contract/audit surface, the annual saving — and state the
    unchanged functional outcomes as fact, not as a disappointment."""
    eid = client.post("/api/engagements", json={"customer_name": "Consolidation Co"}).json()["id"]
    kw = client.post(f"/api/engagements/{eid}/personas",
                     json={"name": "Knowledge Workers", "headcount": 1000}).json()
    client.post(f"/api/engagements/{eid}/current-licenses",
                json={"sku_reference": "Microsoft 365 E5", "quantity_assigned": 1000,
                      "unit_price_paid_annual": 700, "persona_ids": [kw["id"]]})
    # A tool whose capability E5 already delivers, tagged to the persona: the move
    # consolidates it away without lighting up anything new.
    tp = client.post(f"/api/engagements/{eid}/third-party",
                     json={"name": "Acme MFA", "raw_cost": 60000, "cost_period": "Annual",
                           "covered_count": 1000, "persona_ids": [kw["id"]]}).json()
    outcomes = client.get(f"/api/engagements/{eid}/outcomes").json()
    mfa = next(o for o in outcomes if "MFA" in o["name"])
    client.post(f"/api/engagements/{eid}/coverage", json={
        "product_kind": "ThirdParty", "third_party_product_id": tp["id"],
        "outcome_id": mfa["id"], "coverage": "Full", "ratified": True})
    client.post(f"/api/engagements/{eid}/scenarios",
                json={"persona_id": kw["id"], "target_sku_reference": "Microsoft 365 E5",
                      "target_unit_price_annual": 600, "in_scope": True})

    result = client.post(f"/api/engagements/{eid}/compute").json()
    entry = _entry(result, kw["id"])
    assert entry["outcomes"] == []
    assert entry["empty_reason"] == "covered_today"
    text = entry["empty_reason_text"]
    assert "No new functional outcomes" in text          # honest about capability
    assert "Acme MFA" in text                            # names what consolidates
    assert "audit" in text                               # the surface it shrinks
    assert "/yr less" in text                            # and the money it frees
    assert entry["empty_reason_customer_text"] == text   # same story to the customer

    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "Acme MFA" in html.split("<h2>New outcomes</h2>")[1].split("</section>")[0]


def test_unmapped_target_is_reported_as_a_data_gap_not_a_capability_wipe(client):
    """A target with no ratified coverage can neither add nor drop anything —
    the old behaviour showed no new outcomes AND dropped the persona's whole
    current capability, a phantom trade-off on the readout."""
    eid, pid = _engagement(client, "Unmapped Target Co", target="Microsoft 365 E5 (bespoke)")
    result = client.post(f"/api/engagements/{eid}/compute").json()
    entry = _entry(result, pid)
    assert entry["outcomes"] == []
    assert entry["empty_reason"] == "target_unmapped"
    assert "Microsoft 365 E5 (bespoke)" in entry["empty_reason_text"]
    assert "Coverage map" in entry["empty_reason_text"] or "Staple bundles" in entry["empty_reason_text"]
    # No phantom drop, and no Capability trade-offs section built on one.
    assert result["dropped_capability"] == []
    # The customer readout says the comparison is pending — not "map the SKU in
    # Settings", which is an instruction to the operator, not to the customer.
    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "still being confirmed" in html
    assert "Staple bundles" not in html
    assert "Capability trade-offs" not in html

    gaps = client.get(f"/api/engagements/{eid}/coverage-gaps").json()
    gap = next(g for g in gaps["personas"] if g["persona_id"] == pid)
    assert gap["target_unmapped"] is True
    assert gap["dropped_outcomes"] == []
    assert [u["reference"] for u in gap["unmapped_target"]] == ["Microsoft 365 E5 (bespoke)"]


def test_untagged_current_licensing_is_named_as_the_reason(client):
    """An untagged licence line counts for EVERY persona (deliberately
    conservative), so it can make a target look redundant. Say so — with the
    line named — instead of printing nothing."""
    eid, pid = _engagement(client, "Untagged Co", target="Microsoft 365 E3",
                           tag_licence=False)
    result = client.post(f"/api/engagements/{eid}/compute").json()
    entry = _entry(result, pid)
    assert entry["outcomes"] == []
    assert entry["empty_reason"] == "covered_org_wide"
    assert "Microsoft 365 E5" in entry["empty_reason_text"]
    assert "Current licensing tab" in entry["empty_reason_text"]  # the operator's fix

    # The customer-facing readout gets the value story, never the internal fix-it.
    html = client.get(f"/api/engagements/{eid}/readout.html").text
    assert "no persona tag" not in html
    assert "Current licensing tab" not in html
    assert "No new functional outcomes" in html

    gaps = client.get(f"/api/engagements/{eid}/coverage-gaps").json()
    gap = next(g for g in gaps["personas"] if g["persona_id"] == pid)
    assert gap["org_wide_current_licenses"] == ["Microsoft 365 E5"]


def test_tagged_licensing_does_not_raise_the_org_wide_guard(client):
    """The guard fires on untagged lines only — no false alarm on a tagged one."""
    eid, pid = _engagement(client, "Tagged Co", target="Microsoft 365 E3")
    client.post(f"/api/engagements/{eid}/compute")
    gaps = client.get(f"/api/engagements/{eid}/coverage-gaps").json()
    gap = next(g for g in gaps["personas"] if g["persona_id"] == pid)
    assert gap["org_wide_current_licenses"] == []
    assert gap["target_unmapped"] is False


def test_xlsx_accounts_for_a_persona_that_gains_nothing(client):
    eid, _ = _engagement(client, "Xlsx Nothing New Co", target="Microsoft 365 E5")
    data = client.get(f"/api/engagements/{eid}/readout.xlsx").content
    rows = list(load_workbook(io.BytesIO(data))["Capability changes"].iter_rows(values_only=True))
    none_rows = [r for r in rows[1:] if r[2] == "None"]
    assert len(none_rows) == 1
    assert none_rows[0][0] == "Knowledge Workers"
    assert "No new functional outcomes" in none_rows[0][3]


def test_gaining_persona_carries_no_reason(client):
    eid, pid = _engagement(client, "Gain Co", target="Microsoft 365 E5")
    # Swap the persona onto a smaller current licence so the target adds capability.
    lics = client.get(f"/api/engagements/{eid}/current-licenses").json()
    client.patch(f"/api/engagements/{eid}/current-licenses/{lics[0]['id']}",
                 json={"sku_reference": "Microsoft 365 E3"})
    result = client.post(f"/api/engagements/{eid}/compute").json()
    entry = _entry(result, pid)
    assert entry["outcomes"]
    assert entry["empty_reason"] is None


def test_narrative_payload_carries_the_consolidation_note(client):
    """The AI narrative writes `whats_new` per persona. With no new outcomes it
    must have the consolidation story to write from, or it either invents
    capability or writes the persona off — both dishonest."""
    from app.services import narrative
    from app.db import SessionLocal
    from app import models

    eid, pid = _engagement(client, "Narrative Note Co", target="Microsoft 365 E5")
    result = client.post(f"/api/engagements/{eid}/compute").json()
    with SessionLocal() as db:
        payload = narrative.build_narrative_payload(db.get(models.Engagement, eid), result)
    entry = next(p for p in payload if p["persona"] == "Knowledge Workers")
    assert entry["new_outcomes"] == []
    assert "No new functional outcomes" in entry["no_new_outcomes_note"]


def test_narrative_payload_has_no_note_when_the_move_adds_capability(client):
    eid, pid = _engagement(client, "Narrative Gain Co", target="Microsoft 365 E5")
    lics = client.get(f"/api/engagements/{eid}/current-licenses").json()
    client.patch(f"/api/engagements/{eid}/current-licenses/{lics[0]['id']}",
                 json={"sku_reference": "Microsoft 365 E3"})
    result = client.post(f"/api/engagements/{eid}/compute").json()
    from app.services import narrative
    from app.db import SessionLocal
    from app import models
    with SessionLocal() as db:
        payload = narrative.build_narrative_payload(db.get(models.Engagement, eid), result)
    entry = next(p for p in payload if p["persona"] == "Knowledge Workers")
    assert entry["new_outcomes"]
    assert entry["no_new_outcomes_note"] == ""
