"""Readout export (PRD Section 11.5): structured HTML + spreadsheet.

The HTML readout is a customer-facing document: it leads with the headline delta
and (when available) the per-persona business case, then the spend bridge and
scenarios. Internal/QA data (the population check) is not shown, and sections are
conditional — the third-party dispositions, the "what this retires" call-outs, and
the tooling-split / forced-elimination notes appear only when they apply, rather
than being printed as "None". The xlsx export keeps the full QA detail (including
the population check). Deck/one-pager generation is a later AI-assisted feature.
"""

from __future__ import annotations

import html
import io
import re
from decimal import Decimal

from openpyxl import Workbook

from .. import models

# A CSS color we're willing to inline into the readout <style>: a hex triplet or
# an rgb()/rgba() with only digits, commas, dots, spaces. Anything else (which
# could try to break out of the style context) falls back to the default.
_COLOR_RE = re.compile(r"^#[0-9a-fA-F]{3,8}$|^rgba?\([0-9.,\s]+\)$")


def _safe_color(value: str, default: str) -> str:
    value = (value or "").strip()
    return value if _COLOR_RE.match(value) else default


def _is_data_image(value: str) -> bool:
    """True only for a base64-encoded image data URL, so the logo can't inject a
    javascript: or external URL into the readout."""
    return bool(re.match(r"^data:image/[a-zA-Z0-9.+-]+;base64,[A-Za-z0-9+/=\s]+$", value or ""))


def _usd(value) -> str:
    v = float(value or 0)
    return f"−${abs(v):,.2f}" if v < 0 else f"${v:,.2f}"


def _usd0(value) -> str:
    """Compact unsigned money for headline altitude: $246,560. Finance readers
    stall on a minus sign next to the word "savings", so headline surfaces say
    the direction in words ("saved" / "adds") around an unsigned figure; the
    bridge tables keep signed accounting math."""
    return f"${abs(float(value or 0)):,.0f}"


def _pct(value) -> str:
    return f"{float(value or 0) * 100:.0f}%"


def build_html(engagement: models.Engagement, result: dict) -> str:
    """Customer-facing HTML readout. Internal/QA data (the population check) is not
    shown, and sections that don't apply are omitted rather than printed as "None":
    the tooling-split note only when a managed tool exists, the elimination and
    forced-elimination call-outs only when there is something to eliminate, and the
    third-party sections only when third-party tools are in play."""
    rollup = result["rollup"]
    delta = rollup["net_tco_delta_annual"]
    # Cost-change convention: negative = saving (good, green), positive = a cost
    # increase (neutral — spending more isn't an error, just shown honestly).
    delta_cls = "pos" if delta < 0 else ""

    dispositions = result["dispositions"]
    has_third_party = bool(dispositions)
    managed_any = any(d["is_managed"] for d in dispositions)

    rows_scenarios = "".join(
        f"<tr><td>{html.escape(s['persona_name'])}</td>"
        f"<td>{html.escape(s['target_sku_reference'])}</td>"
        f"<td>{s['headcount']}</td>"
        f"<td class='num'>{_usd(s['current_spend_annual'])}</td>"
        f"<td class='num'>{_usd(s['target_spend_annual'])}</td>"
        f"<td class='num {'pos' if s['delta_annual'] < 0 else ''}'>{_usd(s['delta_annual'])}</td>"
        f"<td>{'In scope' if s['in_scope'] else 'Excluded'}</td></tr>"
        for s in result["scenarios"]
    )

    # Third-party dispositions — customer language, never engine enums, and only
    # rows with real dollars behind them (a costless tool is noise, not a finding).
    _action = {
        "FullyEliminated": "Retire fully",
        "PartiallyReduced": "Reduce seats",
        "Unchanged": "Keep",
    }
    disp_rows_src = [
        d for d in dispositions
        if (d.get("effective_annual_cost") or 0) or (d.get("residual_annual_cost") or 0)
    ]
    disp_section = ""
    if disp_rows_src:
        rows_disp = "".join(
            f"<tr><td>{html.escape(d['third_party_product_name'])}</td>"
            f"<td>{_action.get(d['disposition'], d['disposition'])}</td>"
            f"<td class='num'>{min(d['displaced_users'], d['covered_count'])}</td>"
            f"<td class='num'>{d['residual_count']}</td>"
            f"<td class='num'>{_usd(d['residual_annual_cost'])}</td>"
            f"<td>{'Managed service (' + _pct(d['tooling_pct']) + ' tooling share counted)' if d['is_managed'] else 'Direct license'}</td>"
            f"<td>{html.escape(d['override_reason']) if d['override'] != 'None' else ''}</td></tr>"
            for d in disp_rows_src
        )
        disp_section = (
            "<section><h2>Third-party tools — what happens to each</h2>"
            "<table><thead><tr><th>Product</th><th>Action</th><th>Seats retired</th>"
            "<th>Seats kept</th><th>Remaining cost/yr</th><th>Cost basis</th><th>Note</th></tr></thead>"
            f"<tbody>{rows_disp}</tbody></table></section>"
        )

    # Quick wins: third-party duplicates the CURRENT licensing already covers.
    outcome_name = {o.id: o.name for o in engagement.outcomes}
    quick_wins = rollup.get("quick_wins", []) or []
    quick_win_total = rollup.get("quick_win_savings_annual", 0) or 0
    quick_win_rows = "".join(
        f"<tr><td>{html.escape(q['third_party_product_name'])}</td>"
        f"<td>{html.escape(', '.join(outcome_name.get(i, i) for i in q['duplicated_outcome_ids']))}</td>"
        f"<td class='num'>{q['covered_count']}</td>"
        f"<td class='num'>{q['displaced_today']}"
        + (f" ({q['residual_today']} left)" if q["residual_today"] else "")
        + f"</td><td class='num pos'>{_usd(q['credited_annual'])}</td></tr>"
        for q in quick_wins
    )
    quick_win_section = (
        f"<section><h2>Quick wins — you're already covered</h2>"
        f"<p class='sub'>These third-party tools deliver outcomes your <b>current</b> Microsoft "
        f"licensing already provides. They can be retired <b>today</b>, with no licensing "
        f"change: <b class='pos'>{_usd(quick_win_total)}</b>/yr stops immediately.</p>"
        f"<table><thead><tr><th>Tool</th><th>Duplicated capability (already in current licensing)</th>"
        f"<th>Seats on the tool</th><th>Redundant today</th><th>Stops/yr</th></tr></thead>"
        f"<tbody>{quick_win_rows}</tbody></table></section>"
        if quick_wins else ""
    )

    # New outcomes: per in-scope persona, the capabilities the move lights up
    # that nothing they hold today delivers (computed by services/compute from
    # the ratified coverage map — the same source the Coverage Check validates).
    # Rendered as a tile grid — each capability named AND described in plain
    # English, tagged NEW, under a "Persona (headcount) → target" header — so
    # the customer sees what they're getting, not just a list of labels.
    # Omitted entirely when there is nothing new.
    new_outcomes = result.get("new_outcomes") or []
    target_by_pid = {
        s["persona_id"]: s["target_sku_reference"]
        for s in result.get("scenarios", []) if s.get("in_scope")
    }

    def _outcome_chip(o):
        # One outlined chip per capability — scannable, like the GUI's pills;
        # the description rides as hover text (and stays editable data).
        desc = (o.get("description") or "").strip()
        title = f" title=\"{html.escape(desc, quote=True)}\"" if desc else ""
        return f"<span class='chip'{title}>{html.escape(o['name'])}</span>"

    new_outcome_blocks = "".join(
        f"<div class='persona-outcomes'><h3>{html.escape(n['persona_name'])} "
        f"<span class='muted'>({n['headcount']}) → "
        f"{html.escape(target_by_pid.get(n['persona_id'], ''))}</span></h3>"
        f"<div class='chip-row'>{''.join(_outcome_chip(o) for o in n['outcomes'])}</div></div>"
        for n in new_outcomes
    )
    new_outcomes_section = (
        "<section><h2>New outcomes</h2>"
        "<p class='sub'>Capabilities each persona gains with the target licensing that "
        "nothing they hold today delivers — the value the move adds beyond the cost story.</p>"
        f"{new_outcome_blocks}</section>"
        if new_outcome_blocks else ""
    )

    # Spend bridge (cost-change framing): the new target Microsoft cost, minus the
    # existing spend it retires (current Microsoft + freed-up third-party), builds
    # to the net change. The freed third-party splits into "already covered by
    # current licensing (quick win)" and "additionally freed by the move". Every
    # line is broken down per persona: one column per in-scope scenario plus a
    # Total, sourced from the same per-scenario numbers the rollup totals sum
    # (target/current spend and per-product offsets), so columns add up exactly.
    # With a single in-scope persona the total IS that persona — no extra columns.
    existing_ms = rollup["existing_microsoft_annual"]
    target_ms = rollup["target_microsoft_annual"]
    freed = rollup["freed_third_party"]
    already = [f for f in freed if f.get("already_covered")]
    newly = [f for f in freed if not f.get("already_covered")]
    already_ids = {f["third_party_product_id"] for f in already}
    in_scope = [s for s in result["scenarios"] if s["in_scope"]]
    cols = in_scope if len(in_scope) > 1 else []

    def _offset_field(s, product_id, field):
        return next((o.get(field, 0) for o in s.get("offsets", [])
                     if o["third_party_product_id"] == product_id), 0)

    def _fmt(value, negate):
        return (f"−{_usd(value)}" if value else _usd(0)) if negate else _usd(value)

    def _cells(values, total, negate=False, cls=""):
        tds = "".join(f"<td class='num {cls}'>{_fmt(v, negate)}</td>" for v in values)
        return tds + f"<td class='num {cls}'>{_fmt(total, negate)}</td>"

    bridge_head = ""
    if cols:
        ths = "".join(
            f"<th class='num'>{html.escape(s['persona_name'])} "
            f"<small>→ {html.escape(s['target_sku_reference'])}</small></th>"
            for s in cols
        )
        bridge_head = f"<thead><tr><th></th>{ths}<th class='num'>Total</th></tr></thead>"

    def _freed_group(label, sub, items, field):
        """One bridge group: products contributing via `field` (dollar rows of
        zero are dropped — no placeholder rows on a customer document)."""
        rows_items = [f for f in items if f.get(field, 0)]
        if not rows_items:
            return ""
        total = sum(f[field] for f in rows_items)
        rows = "".join(
            f"<tr class='sub'><td>↳ {html.escape(f['third_party_product_name'])}</td>"
            + _cells([_offset_field(s, f["third_party_product_id"],
                                    {"redundant_today_annual": "redundant_today_annual",
                                     "move_unlocked_annual": "move_unlocked_annual",
                                     "credited_annual": "credited_offset_annual"}[field])
                      for s in cols],
                     f[field], negate=True, cls="pos")
            + "</tr>"
            for f in rows_items
        )
        group_total_cells = _cells(
            [sum(_offset_field(s, f["third_party_product_id"],
                               {"redundant_today_annual": "redundant_today_annual",
                                "move_unlocked_annual": "move_unlocked_annual",
                                "credited_annual": "credited_offset_annual"}[field])
                 for f in rows_items) for s in cols],
            total, negate=True, cls="pos")
        return (f"<tr><td>Less: {label} <span class='muted'>{sub}</span></td>"
                f"{group_total_cells}</tr>{rows}")

    delta_word = (
        f"{_usd0(delta)}/yr saved" if delta < 0
        else f"{_usd0(delta)}/yr added" if delta > 0 else "no net change"
    )
    delta_cells = "".join(
        f"<td class='num {'pos' if s['delta_annual'] < 0 else ''}'><b>{_usd(s['delta_annual'])}</b></td>"
        for s in cols
    )
    bridge_rows = (
        f"<tr><td>Target Microsoft licensing (new per-persona bundles)</td>"
        + _cells([s["target_spend_annual"] for s in cols], target_ms)
        + "</tr>"
        f"<tr><td>Less: existing Microsoft licensing retired (current assigned)</td>"
        + _cells([s["current_microsoft_annual"] for s in cols], existing_ms,
                 negate=True, cls="pos")
        + "</tr>"
        + _freed_group("third-party redundant today",
                       "(the quick wins — no licensing change required)",
                       already, "redundant_today_annual")
        + _freed_group("further seats on those tools retired by the move",
                       "(covered only once the target lands)",
                       already, "move_unlocked_annual")
        + _freed_group("third-party newly retired by the move",
                       "(capability arrives with the target)",
                       newly, "credited_annual")
        + f"<tr class='total'><td><b>Net change in run-rate</b> "
        f"<span class='muted'>({delta_word})</span></td>"
        f"{delta_cells}<td class='num {delta_cls}'><b>{_usd(delta)}</b></td></tr>"
    )
    # Reconciliation note: the "redundant today" group equals the Quick Wins
    # total whenever every quick-win tool is displaced by a move. When one
    # isn't (its savings live outside this bridge), say so explicitly instead
    # of leaving two sections that disagree.
    bridged_today = float(rollup.get("freed_redundant_today_annual", 0) or 0)
    qw_check = float(rollup.get("quick_win_savings_annual", 0) or 0)
    bridge_note = ""
    if abs(bridged_today - qw_check) >= 0.005:
        bridge_note = (
            f"<p class='sub'>A further {_usd(qw_check - bridged_today)}/yr of quick-win "
            f"savings sits outside this bridge — tools retirable today that no persona "
            f"move touches. See Quick wins above; the two together make the total "
            f"opportunity in the headline.</p>"
        )

    # Hero block (Section 6.8a decomposition): the headline splits into the two
    # things the customer can actually decide — ① retire duplicate tools today
    # (the quick wins, no licensing change) and ② move each persona to
    # right-sized licensing (the moves' OWN value, quick-win credit stripped so
    # the two never double-count a dollar). The anchor number is their sum over
    # the modeling horizon, stated in words ("saved"), never as a signed
    # figure next to the word savings.
    horizon = int(engagement.modeling_horizon_years or 3)
    months = horizon * 12
    qw_total = float(rollup.get("quick_win_savings_annual", 0) or 0)
    move_incr = float(rollup.get("move_incremental_delta_annual", 0) or 0)
    moves_value = -move_incr  # positive = the moves save money
    total_opportunity = qw_total + moves_value  # positive = total saved / yr

    if total_opportunity > 0:
        head_word, head_cls = f"saved over {months} months", "pos"
    elif total_opportunity < 0:
        head_word, head_cls = f"added cost over {months} months", ""
    else:
        head_word, head_cls = "no net change", ""

    def _moves_amount(v):
        # Number first, like card ① — the eye scans the amounts down the left.
        if v < 0:
            return f"<span class='move-amt pos'>saves {_usd0(v)}/yr</span>"
        if v > 0:
            return f"<span class='move-amt'>adds {_usd0(v)}/yr</span>"
        return "<span class='move-amt muted'>cost-neutral</span>"

    move_items = "".join(
        f"<li>{_moves_amount(s.get('move_incremental_delta_annual', s['delta_annual']))}"
        f"<b>{html.escape(s['persona_name'])}</b> ({s['headcount']}) → "
        f"<b>{html.escape(s['target_sku_reference'])}</b></li>"
        for s in in_scope
    )
    part_today = (
        f"<div class='hero-part'><div class='part-label'>① Retire duplicate tools today "
        f"— no licensing change</div>"
        f"<div class='part-value pos'>{_usd0(qw_total)}/yr</div></div>"
        if qw_total > 0 else ""
    )
    # Card ② is just the moves — persona → bundle (amount), no roll-up number
    # repeating what the bridge derives later.
    part_moves = (
        f"<div class='hero-part'><div class='part-label'>"
        f"{'② ' if part_today else ''}Move each persona to right-sized licensing</div>"
        + (f"<ul class='moves'>{move_items}</ul>" if in_scope else "")
        + "</div>"
    )
    # List-price caveat: when baseline spend rests on assumed prices, say so
    # next to the headline, not in the appendix — it builds trust, not doubt.
    assumed_n = sum(1 for lic in engagement.current_licenses if lic.source_tag == "ListPrice")
    hero_caveat = (
        f"<div class='hero-caveat'>Baseline spend uses list-price assumptions for "
        f"{assumed_n} current SKU{'s' if assumed_n != 1 else ''} — validating against "
        f"invoices is step one and may move this number. Figures are run-rate: they "
        f"assume retirements from day one; year one phases with contract end dates.</div>"
        if assumed_n
        else "<div class='hero-caveat'>Figures are run-rate: they assume retirements "
             "from day one; year one phases with contract end dates.</div>"
    )
    # One headline, two stacked sub-cards. The components' dollars live in the
    # cards ONLY — no equation line repeating them above.
    hero_sub = f"{_usd0(total_opportunity)}/yr run-rate"
    hero = (
        f"<section class='hero'>"
        f"<div class='hero-label'>Total opportunity <span class='hero-note'>"
        f"· {horizon}-year run-rate view · quick wins + licensing moves</span></div>"
        f"<div class='headline {head_cls}'>{_usd0(total_opportunity * horizon)} "
        f"<span class='headline-word'>{head_word}</span></div>"
        f"<div class='hero-sub'>{hero_sub}</div>"
        f"<div class='hero-split'>{part_today}{part_moves}</div>"
        f"{hero_caveat}"
        f"</section>"
    )

    # Eliminations section — build only the parts that have content, and omit the
    # whole section if nothing was eliminated (don't print "None" to a customer).
    elim_parts = []
    fully_elim = rollup["fully_eliminated_tools"]
    renewal_by_name = {
        r["third_party_product_name"]: r["renewal_date"]
        for r in rollup["eliminated_renewal_cycles"]
        if r.get("renewal_date")
    }
    if fully_elim:
        items = "".join(
            f"<li>{html.escape(t)}"
            + (f" <span class='muted'>— contract renews {renewal_by_name[t]}; "
               f"co-term the retirement against this date</span>"
               if t in renewal_by_name else "")
            + "</li>"
            for t in fully_elim
        )
        elim_parts.append(f"<p><b>Tools fully eliminated:</b></p><ul>{items}</ul>")
    if has_third_party:
        elim_parts.append(
            f"<p><b>Third-party spend that remains after the move:</b> "
            f"{_usd(rollup['residual_third_party_cost_annual'])}/yr "
            f"<span class='muted'>(reduced seats plus tools kept as-is)</span></p>"
        )
    elim_section = (
        f"<section><h2>What this retires</h2>{''.join(elim_parts)}</section>"
        if elim_parts else ""
    )

    # The business case (per-persona narrative). Advisory, generated elsewhere and
    # attached to the result when available; the section is omitted when it isn't.
    narratives = result.get("narratives") or []
    narrative_section = ""
    if narratives:
        blocks = "".join(
            f"<div class='narrative'><h3>{html.escape(n.get('persona', ''))}</h3>"
            + (f"<p><b>Today:</b> {html.escape(n.get('today', ''))}</p>" if n.get("today") else "")
            + (f"<p><b>What's new:</b> {html.escape(n.get('whats_new', ''))}</p>" if n.get("whats_new") else "")
            + (f"<p><b>Value:</b> {html.escape(n.get('value', ''))}</p>" if n.get("value") else "")
            + "</div>"
            for n in narratives
        )
        narrative_section = f"<section><h2>The business case</h2>{blocks}</section>"

    # Assumptions & sources — only the notes that apply to this engagement.
    appendix_parts = []
    # The from-state anchors the whole story: show what the customer holds
    # today (the current-licensing mix), never leave it invisible.
    if engagement.current_licenses:
        lic_rows = "".join(
            f"<tr><td>{html.escape(lic.sku_reference)}"
            + (" <span class='muted'>(list price assumed)</span>"
               if lic.source_tag == "ListPrice" else "")
            + f"</td><td class='num'>{lic.quantity_assigned}</td>"
            f"<td class='num'>{_usd(lic.unit_price_paid_annual)}</td>"
            f"<td class='num'>{_usd(float(lic.unit_price_paid_annual or 0) * (lic.quantity_assigned or 0))}</td></tr>"
            for lic in engagement.current_licenses
        )
        appendix_parts.append(
            "<p><b>Current Microsoft licensing (as provided):</b></p>"
            "<table><thead><tr><th>SKU</th><th>Seats assigned</th>"
            "<th>Price paid /seat/yr</th><th>Annual</th></tr></thead>"
            f"<tbody>{lic_rows}</tbody></table>"
        )
    # Target pricing basis — the first procurement question in the room.
    _term_label = {"P1Y": "1-year term", "P1M": "monthly term", "P3Y": "3-year term"}
    appendix_parts.append(
        f"<p><b>Target pricing basis:</b> {html.escape(engagement.default_segment or 'Commercial')} · "
        f"{html.escape(_term_label.get(engagement.default_term_duration or 'P1Y', engagement.default_term_duration or 'P1Y'))} · "
        f"{html.escape(engagement.default_billing_plan or 'Monthly')} billing — catalog price "
        f"unless an operator-entered price overrides a line.</p>"
    )
    if managed_any:
        tooling_overrides = [
            d for d in dispositions
            if d["is_managed"] and abs(d["tooling_pct"] - float(engagement.global_tooling_pct)) > 1e-9
        ]
        override_html = ""
        if tooling_overrides:
            items = "".join(
                f"<li>{html.escape(d['third_party_product_name'])}: {_pct(d['tooling_pct'])}</li>"
                for d in tooling_overrides
            )
            override_html = f"<p>Per-line overrides:</p><ul>{items}</ul>"
        appendix_parts.append(
            "<p><b>Tooling split.</b> Managed third-party products count at their tooling "
            "percentage only; management of M365 is presumed neutral or better. Engagement "
            f"default: {_pct(engagement.global_tooling_pct)}.</p>{override_html}"
        )
    overrides = [d for d in dispositions if d["override"] == "ForceFullElimination"]
    if overrides:
        items = "".join(
            f"<li><b>{html.escape(d['third_party_product_name'])}</b>: {html.escape(d['override_reason'])}</li>"
            for d in overrides
        )
        appendix_parts.append(
            "<p><b>Assumed full elimination</b> (savings asserted on users the target does "
            f"not automatically displace):</p><ul>{items}</ul>"
        )
    # Only lines with a recorded discount are disclosed — an appendix line must
    # say something real, never print a placeholder to the customer.
    discounted = [lic for lic in engagement.current_licenses if lic.discount_pct]
    if discounted:
        items = "".join(
            f"<li>{html.escape(lic.sku_reference)}: {_pct(lic.discount_pct)} discount</li>"
            for lic in discounted
        )
        appendix_parts.append(f"<p><b>Current Microsoft line discounts:</b></p><ul>{items}</ul>")
    # Input provenance (DATA_MODEL §9): separate hard inputs from soft ones.
    # Disclose every input whose source_tag marks it as an assumption rather than
    # a customer-stated/invoiced fact; omitted entirely when there are none.
    soft_label = {
        "ListPrice": "list price assumed",
        "Estimate": "estimate",
        "AISuggestedUnconfirmed": "AI-suggested, unconfirmed",
    }
    soft_inputs = (
        [(p.name, "persona", p.source_tag) for p in engagement.personas
         if p.source_tag in soft_label]
        + [(lic.sku_reference, "current license", lic.source_tag)
           for lic in engagement.current_licenses if lic.source_tag in soft_label]
        + [(tp.name, "third-party product", tp.source_tag)
           for tp in engagement.third_party_products if tp.source_tag in soft_label]
    )
    if soft_inputs:
        items = "".join(
            f"<li>{html.escape(name)} <span class='muted'>({kind})</span>: "
            f"{soft_label[tag]}</li>"
            for name, kind, tag in soft_inputs
        )
        appendix_parts.append(
            "<p><b>Inputs carried as assumptions</b> (tagged, not customer-stated "
            f"or invoiced):</p><ul>{items}</ul>"
        )
    appendix_section = (
        f"<section><h2>Assumptions &amp; sources</h2>{''.join(appendix_parts)}</section>"
        if appendix_parts else ""
    )

    # The "then what": a readout that ends at a footer leaves no path from
    # assumed to confirmed. Validate → Sequence → Fund → Decide, plus the two
    # objections a CIO raises unprompted (change management, one-time cost).
    n_retire = len(fully_elim)
    next_steps_section = (
        "<section><h2>Recommended next steps</h2><ol>"
        "<li><b>Validate.</b> Confirm current licensing counts and third-party invoices, "
        f"and pull contract end dates for the {n_retire} retirement target{'s' if n_retire != 1 else ''} "
        "— this converts the assumptions above into an invoice-verified business case.</li>"
        "<li><b>Sequence.</b> Co-term each retirement against its renewal date into a "
        "phased savings schedule. The figures here are run-rate; year one phases in.</li>"
        "<li><b>Fund.</b> Microsoft co-investment programs may offset transition cost — "
        "quantified in the follow-on.</li>"
        "<li><b>Decide.</b> A 30-day validation sprint converts this readout into an "
        "invoice-verified, phased business case.</li></ol>"
        "<p class='sub'>Noted up front: savings are gross of one-time migration cost, and "
        "workload moves (chat, device management, identity) are change-management events, "
        "not just license swaps — both are scoped in the follow-on.</p></section>"
        if in_scope else ""
    )

    # Readout branding (user-entered). Sanitize hard: colors must match a strict
    # CSS-color pattern and the logo must be a base64 image data URL, so neither
    # can break out of the <style>/<img> context. Blank -> neutral built-in theme.
    primary = _safe_color(engagement.brand_primary_color, "#1a1a2e")
    accent = _safe_color(engagement.brand_accent_color, "#2563eb")
    logo = engagement.brand_logo_data_url or ""
    logo_html = (
        f'<img src="{html.escape(logo, quote=True)}" alt="logo" '
        f'style="max-height:56px;max-width:240px;margin-bottom:.5rem">'
        if _is_data_image(logo) else ""
    )

    return f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>M365 TCO Readout — {html.escape(engagement.customer_name)}</title>
<style>
 :root{{--primary:{primary};--accent:{accent};--pos:#127436;--neg:#b00020;
   --ink:#1f2430;--muted:#5b6472;--line:#e5e8ee;--soft:#f6f8fb}}
 *{{box-sizing:border-box}}
 body{{font-family:-apple-system,'Segoe UI',Roboto,'Helvetica Neue',sans-serif;
   margin:0;color:var(--ink);background:#fff;line-height:1.45}}
 main{{max-width:1020px;margin:0 auto;padding:2.2rem 2rem 3rem}}
 h1{{margin:.2rem 0 0;font-size:1.65rem;color:var(--primary);letter-spacing:-.01em}}
 .sub{{color:var(--muted);font-size:.94rem}}
 .hero{{margin:1.5rem 0 2rem;padding:1.3rem 1.6rem;background:var(--soft);
   border:1px solid var(--line);border-left:4px solid var(--accent);border-radius:10px}}
 .hero-label{{font-size:.8rem;font-weight:650;text-transform:uppercase;
   letter-spacing:.08em;color:var(--muted)}}
 .hero-note{{font-weight:400;text-transform:none;letter-spacing:0}}
 .headline{{font-size:3rem;font-weight:750;line-height:1.1;margin:.15rem 0;letter-spacing:-.02em}}
 .hero-sub{{color:var(--muted)}}
 .headline-word{{font-size:1.15rem;font-weight:600;color:var(--muted);letter-spacing:0}}
 .hero-split{{display:flex;flex-direction:column;gap:.7rem;margin-top:.9rem;
   padding-top:.9rem;border-top:1px solid var(--line)}}
 .hero-part{{background:#fff;border:1px solid var(--line);
   border-radius:8px;padding:.6rem .85rem}}
 .part-label{{font-size:.8rem;font-weight:650;color:var(--muted)}}
 .part-value{{font-size:1.35rem;font-weight:750;margin:.1rem 0}}
 .hero-caveat{{margin-top:.7rem;font-size:.82rem;color:var(--muted)}}
 ul.moves{{list-style:none;margin:.35rem 0 0;padding:0}}
 ul.moves li{{margin:.25rem 0;font-size:.95rem}}
 .move-amt{{display:inline-block;min-width:10.5rem;font-weight:700}}
 .pos{{color:var(--pos)}} .neg{{color:var(--neg)}} .muted{{color:var(--muted)}}
 section{{margin:2rem 0}}
 h2{{font-size:1.12rem;color:var(--primary);margin:0 0 .4rem;
   padding-bottom:.3rem;border-bottom:1px solid var(--line)}}
 table{{border-collapse:collapse;width:100%;margin:.8rem 0;font-size:.92rem}}
 th{{color:var(--muted);font-weight:600;font-size:.76rem;text-transform:uppercase;
   letter-spacing:.05em;text-align:left;padding:.5rem .65rem;border-bottom:2px solid var(--line)}}
 td{{padding:.5rem .65rem;border-bottom:1px solid var(--line);text-align:left}}
 td.num,th.num{{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}}
 table.bridge td{{border:none;padding:.32rem .65rem}}
 table.bridge th{{border:none;text-transform:none;letter-spacing:0;font-size:.85rem;padding:.32rem .65rem}}
 table.bridge tr.sub td{{color:var(--muted);padding-left:1.9rem;font-size:.85rem}}
 table.bridge tr.total td{{border-top:2px solid var(--ink);padding-top:.5rem}}
 .narrative{{background:var(--soft);border-left:3px solid var(--accent);
   padding:.6rem 1rem;border-radius:8px;margin:.75rem 0}}
 .narrative h3{{margin:.2rem 0;color:var(--primary)}}
 .persona-outcomes{{background:var(--soft);border:1px solid var(--line);
   border-left:3px solid var(--pos);border-radius:10px;padding:.7rem .95rem;margin:.7rem 0}}
 .persona-outcomes h3{{margin:0 0 .45rem;font-size:1rem;color:var(--primary)}}
 .chip-row{{display:flex;flex-wrap:wrap;gap:.4rem}}
 .chip{{display:inline-block;border:1px solid var(--pos);color:var(--pos);
   background:#fff;border-radius:999px;padding:.22rem .7rem;font-size:.82rem;
   font-weight:600;line-height:1.2;white-space:nowrap}}
 ul{{margin:.3rem 0}}
 footer{{margin-top:2.5rem;padding-top:1rem;border-top:1px solid var(--line);
   color:var(--muted);font-size:.8rem}}
 @media (max-width:720px){{
   main{{padding:1.1rem .9rem 2rem}}
   .headline{{font-size:2rem}}
   .headline-word{{display:block;font-size:1rem;margin-top:.1rem}}
   .hero{{padding:.9rem 1rem}}
   table{{display:block;overflow-x:auto;-webkit-overflow-scrolling:touch}}
   td:first-child,th:first-child{{min-width:190px}}
   th,td{{padding:.42rem .5rem}}
 }}
</style></head><body><main>
<header>
{logo_html}
<h1>M365 TCO Readout</h1>
<div class="sub">{html.escape(engagement.customer_name)} · {html.escape(engagement.market or "")}/{html.escape(engagement.currency or "USD")} · annualized {html.escape(engagement.currency or "USD")}</div>
</header>
{hero}
{narrative_section}
{quick_win_section}

<section><h2>Per-persona scenarios</h2>
<table><thead><tr><th>Persona</th><th>Target SKU</th><th>Headcount</th>
<th>Current total spend/yr<br><small>(Microsoft + attributed third-party)</small></th>
<th>Target Microsoft/yr</th><th>Change/yr</th><th>Scope</th></tr></thead>
<tbody>{rows_scenarios}</tbody></table></section>
{new_outcomes_section}

<section><h2>How we get to the number</h2>
<p class="sub">Existing annualized spend for the in-scope population, the third-party
tooling those users free up when they move, and the target Microsoft licensing —
building to the net TCO delta, with each line broken down per persona.</p>
<table class="bridge">{bridge_head}<tbody>{bridge_rows}</tbody></table>{bridge_note}</section>
{disp_section}
{elim_section}
{appendix_section}
{next_steps_section}
<footer>Licensing-only view. Migration services, Microsoft funding, Azure consumption,
and operational savings are quantified in the follow-on. Generated by the M365 TCO Tool.</footer>
</main></body></html>"""


def build_xlsx(engagement: models.Engagement, result: dict) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Scenarios"
    ws.append(
        ["Persona", "Target SKU", "Headcount", "In scope", "Current spend",
         "Target spend", "Delta", "Current MS", "Third-party offset"]
    )
    for s in result["scenarios"]:
        ws.append([
            s["persona_name"], s["target_sku_reference"], s["headcount"],
            s["in_scope"], s["current_spend_annual"], s["target_spend_annual"],
            s["delta_annual"], s["current_microsoft_annual"],
            s["current_third_party_offset_annual"],
        ])

    wd = wb.create_sheet("Dispositions")
    wd.append(
        ["Product", "Disposition", "Covered", "Displaced", "Residual count",
         "Residual cost", "Per-unit", "Effective cost", "Managed", "Tooling %",
         "Override", "Override reason", "Residual intent", "Renewal date"]
    )
    for d in result["dispositions"]:
        wd.append([
            d["third_party_product_name"], d["disposition"], d["covered_count"],
            d["displaced_users"], d["residual_count"], d["residual_annual_cost"],
            d["per_unit_annual_cost"], d["effective_annual_cost"], d["is_managed"],
            d["tooling_pct"], d["override"], d["override_reason"],
            d["residual_intent"], d["renewal_date"],
        ])

    wr = wb.create_sheet("Rollup")
    rollup = result["rollup"]
    pop = rollup["population_check"]
    wr.append(["Metric", "Value"])
    wr.append(["Existing Microsoft licensing (annual, in scope)", rollup["existing_microsoft_annual"]])
    wr.append(["Existing third-party freed up (annual, in scope)", rollup["existing_third_party_annual"]])
    for f in rollup["freed_third_party"]:
        wr.append([f"  freed up: {f['third_party_product_name']}", f["credited_annual"]])
    wr.append([
        "Total existing spend (annual, in scope)",
        rollup["existing_microsoft_annual"] + rollup["existing_third_party_annual"],
    ])
    wr.append(["Target Microsoft licensing (annual, in scope)", rollup["target_microsoft_annual"]])
    wr.append(["Net TCO delta (annual) — negative = saving", rollup["net_tco_delta_annual"]])
    wr.append(["  of which: quick wins (retirable today, no move)",
               -(rollup.get("freed_redundant_today_annual", 0) or 0)])
    wr.append(["  of which: the moves' own value (excl. quick wins)",
               rollup.get("move_incremental_delta_annual", 0) or 0])
    horizon = int(engagement.modeling_horizon_years or 3)
    wr.append([
        f"Net TCO delta ({horizon * 12}-month headline)",
        rollup["net_tco_delta_annual"] * horizon,
    ])
    wr.append(["Residual third-party cost (annual)", rollup["residual_third_party_cost_annual"]])
    wr.append(["In-scope persona headcount", pop["in_scope_persona_headcount"]])
    wr.append(["Third-party covered population", pop["third_party_covered_population"]])
    wr.append(["Fully eliminated tools", ", ".join(rollup["fully_eliminated_tools"])])
    wr.append([
        "Eliminated renewal cycles",
        ", ".join(r["third_party_product_name"] for r in rollup["eliminated_renewal_cycles"]),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
